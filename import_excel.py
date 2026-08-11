"""
اسکریپت یک‌بار اجرا برای ایمپورت *اولیه* اکسل به دیتابیس.
"""
import asyncio
import openpyxl
from database import init_db, clear_products, add_product, count_products

EXCEL_PATH = "/home/workdir/attachments/hamed.xlsx"


async def main():
    await init_db()
    await clear_products()

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    header_row = None
    for row in range(1, 6):
        vals = [ws.cell(row, c).value for c in range(1, 20)]
        if any(v and "دسته" in str(v) for v in vals):
            header_row = row
            break
    if not header_row:
        # fallback: row 1 from the preview
        header_row = 1

    col_map = {}
    for c in range(1, 25):
        val = ws.cell(header_row, c).value
        if not val:
            continue
        v = str(val).strip()
        if "شناسه" in v:
            col_map["excel_id"] = c
        elif "دسته" in v:
            col_map["category"] = c
        elif "خصوصیت 1" in v or "خصوصیت۱" in v:
            col_map["attr1"] = c
        elif "خصوصیت 2" in v or "خصوصیت۲" in v:
            col_map["attr2"] = c
        elif "خصوصیت 3" in v or "خصوصیت۳" in v:
            col_map["attr3"] = c
        elif "خصوصیت 4" in v or "خصوصیت۴" in v:
            col_map["attr4"] = c
        elif "خصوصیت 5" in v or "خصوصیت۵" in v:
            col_map["attr5"] = c
        elif "برند" in v:
            col_map["brand"] = c
        elif "قیمت خرید" in v:
            col_map["purchase_price"] = c
        elif "فروش واحد تک" in v or "قیمت فروش واحد تک" in v:
            col_map["sale_price"] = c
        elif "عمده" in v:
            col_map["wholesale_price"] = c
        elif "فروشنده" in v and "کد" not in v:
            col_map["seller_name"] = c
        elif "کد فروشنده" in v:
            col_map["seller_code"] = c
        elif "بارکد" in v:
            col_map["barcode"] = c
        elif "تاریخ خرید - روز" in v or (v == "تاریخ خرید - روز"):
            col_map["purchase_day"] = c
        elif "تاریخ خرید - ماه" in v:
            col_map["purchase_month"] = c
        elif "تاریخ خرید - سال" in v:
            col_map["purchase_year"] = c

    print("Column map:", col_map)

    count = 0
    for row in range(header_row + 1, ws.max_row + 1):
        cat = ws.cell(row, col_map.get("category", 2)).value
        if not cat or not str(cat).strip():
            continue

        def get(key):
            c = col_map.get(key)
            if not c:
                return None
            v = ws.cell(row, c).value
            if v is None or str(v).strip() == "":
                return None
            return v

        def get_num(key):
            v = get(key)
            if v is None:
                return 0
            try:
                return float(v)
            except Exception:
                return 0

        def get_int(key):
            v = get(key)
            if v is None:
                return None
            try:
                return int(float(v))
            except Exception:
                return None

        excel_id_val = get("excel_id")
        barcode_val = get("barcode")
        data = {
            "excel_id": str(excel_id_val).strip() if excel_id_val is not None else None,
            "category": str(get("category")).strip() if get("category") else None,
            "attr1": str(get("attr1")).strip() if get("attr1") else None,
            "attr2": str(get("attr2")).strip() if get("attr2") else None,
            "attr3": str(get("attr3")).strip() if get("attr3") else None,
            "attr4": str(get("attr4")).strip() if get("attr4") else None,
            "attr5": str(get("attr5")).strip() if get("attr5") else None,
            "brand": str(get("brand")).strip() if get("brand") else None,
            "purchase_price": get_num("purchase_price"),
            "sale_price": get_num("sale_price"),
            "wholesale_price": get_num("wholesale_price"),
            "stock": 0,
            "barcode": str(barcode_val).strip() if barcode_val is not None else None,
            "seller_name": str(get("seller_name")).strip() if get("seller_name") else None,
            "seller_code": str(get("seller_code")).strip() if get("seller_code") else None,
            "photo_file_id": None,
            "purchase_day": get_int("purchase_day"),
            "purchase_month": get_int("purchase_month"),
            "purchase_year": get_int("purchase_year"),
            "sale_percent": None,
            "wholesale_percent": None,
        }
        await add_product(data)
        count += 1

    total = await count_products()
    print(f"Imported {count} products. Total in DB: {total}")


if __name__ == "__main__":
    asyncio.run(main())
