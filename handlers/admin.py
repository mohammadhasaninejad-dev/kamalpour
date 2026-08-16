from aiogram import Router, F
from aiogram.types import Message, CallbackQuery, ContentType, FSInputFile
from aiogram.fsm.context import FSMContext
import openpyxl
import tempfile
import os
import inspect

from states import AddProduct, EditProduct, DeleteProduct, ImportExcel, SellerSearch, AddHistory, IdSearch, EditHistory
from keyboards import (
    admin_menu, main_menu, cancel_kb, skip_kb, barcode_kb, year_suggest_kb,
    confirm_delete_kb, product_actions_kb, edit_fields_kb,
    confirm_wipe_kb, browse_categories_kb, browse_attrs_kb, browse_products_nav_kb,
    results_kb, list_products_menu_kb, history_item_kb, confirm_delete_history_kb
)
from database import (
    add_product, update_product, delete_product, get_product,
    count_products, build_product_name, get_product_by_name,
    get_product_by_excel_id, clear_products,
    get_distinct_categories, get_distinct_attr1,
    get_products_by_category_attr1, count_products_by_category_attr1,
    get_all_products_full, search_by_seller, count_by_seller,
    add_price_history, get_price_history, ensure_initial_history,
    get_history_item, update_price_history, delete_price_history,
    sync_product_from_latest_history, normalize_digits
)
from utils import is_admin, format_product, format_price, parse_number, parse_int, format_date
from config import PAGE_SIZE

router = Router()

CURRENT_YEAR = 1405


def admin_only(handler):
    handler_params = inspect.signature(handler).parameters

    async def wrapper(event, **kwargs):
        user = event.from_user if hasattr(event, "from_user") else None
        if not is_admin(user):
            if isinstance(event, Message):
                await event.answer("⛔ دسترسی فقط برای ادمین.")
            elif isinstance(event, CallbackQuery):
                await event.answer("⛔ دسترسی فقط برای ادمین.", show_alert=True)
            return
        accepted = {k: v for k, v in kwargs.items() if k in handler_params}
        return await handler(event, **accepted)
    return wrapper


# ---------- منوی مدیریت ----------
@router.message(F.text == "📦 مدیریت کالا")
@admin_only
async def admin_panel(message: Message, state: FSMContext):
    await state.clear()
    total = await count_products()
    await message.answer(
        f"پنل مدیریت کالا\nتعداد کل کالاها: {total}",
        reply_markup=admin_menu()
    )


@router.message(F.text == "🔙 بازگشت به مدیریت کالا")
@admin_only
async def back_to_admin(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("پنل مدیریت:", reply_markup=admin_menu())


# ---------- دکمه‌های منوی اصلی ادمین ----------
@router.message(F.text == "📋 لیست کالا بر اساس دسته")
@admin_only
async def main_browse_category(message: Message, state: FSMContext):
    await state.clear()
    categories = await get_distinct_categories()
    if not categories:
        await message.answer("هنوز کالایی ثبت نشده.", reply_markup=main_menu(True))
        return
    await state.update_data(browse_categories=categories)
    await message.answer(
        f"📋 {len(categories)} دسته موجود است. یک دسته را انتخاب کنید:",
        reply_markup=browse_categories_kb(categories, page=0)
    )


@router.message(F.text == "👤 جستجو بر اساس فروشنده")
@admin_only
async def main_seller_search(message: Message, state: FSMContext):
    await state.set_state(SellerSearch.waiting_name)
    await message.answer("نام فروشنده را وارد کنید:", reply_markup=cancel_kb())


# ---------- افزودن کالا ----------
@router.message(F.text == "➕ افزودن کالا")
@admin_only
async def add_start(message: Message, state: FSMContext):
    await state.set_state(AddProduct.category)
    await message.answer("دسته کالا را وارد کنید (مثال: دفتر):", reply_markup=cancel_kb())


@router.message(AddProduct.category, F.text == "❌ انصراف")
@router.message(AddProduct.attr1, F.text == "❌ انصراف")
@router.message(AddProduct.attr2, F.text == "❌ انصراف")
@router.message(AddProduct.attr3, F.text == "❌ انصراف")
@router.message(AddProduct.attr4, F.text == "❌ انصراف")
@router.message(AddProduct.attr5, F.text == "❌ انصراف")
@router.message(AddProduct.brand, F.text == "❌ انصراف")
@router.message(AddProduct.purchase_price, F.text == "❌ انصراف")
@router.message(AddProduct.sale_price, F.text == "❌ انصراف")
@router.message(AddProduct.wholesale_price, F.text == "❌ انصراف")
@router.message(AddProduct.stock, F.text == "❌ انصراف")
@router.message(AddProduct.barcode, F.text == "❌ انصراف")
@router.message(AddProduct.seller_name, F.text == "❌ انصراف")
@router.message(AddProduct.seller_code, F.text == "❌ انصراف")
@router.message(AddProduct.purchase_day, F.text == "❌ انصراف")
@router.message(AddProduct.purchase_month, F.text == "❌ انصراف")
@router.message(AddProduct.purchase_year, F.text == "❌ انصراف")
@router.message(AddProduct.photo, F.text == "❌ انصراف")
@router.message(AddHistory.day, F.text == "❌ انصراف")
@router.message(AddHistory.month, F.text == "❌ انصراف")
@router.message(AddHistory.year, F.text == "❌ انصراف")
@router.message(AddHistory.price, F.text == "❌ انصراف")
async def add_cancel(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("لغو شد.", reply_markup=main_menu(is_admin(message.from_user)))


@router.message(AddProduct.category)
async def add_category(message: Message, state: FSMContext):
    await state.update_data(category=message.text.strip())
    await state.set_state(AddProduct.attr1)
    await message.answer("خصوصیت ۱ را وارد کنید (یا رد کردن):", reply_markup=skip_kb())


@router.message(AddProduct.attr1)
async def add_attr1(message: Message, state: FSMContext):
    val = None if message.text == "⏭ رد کردن" else message.text.strip()
    await state.update_data(attr1=val)
    await state.set_state(AddProduct.attr2)
    await message.answer("خصوصیت ۲:", reply_markup=skip_kb())


@router.message(AddProduct.attr2)
async def add_attr2(message: Message, state: FSMContext):
    val = None if message.text == "⏭ رد کردن" else message.text.strip()
    await state.update_data(attr2=val)
    await state.set_state(AddProduct.attr3)
    await message.answer("خصوصیت ۳:", reply_markup=skip_kb())


@router.message(AddProduct.attr3)
async def add_attr3(message: Message, state: FSMContext):
    val = None if message.text == "⏭ رد کردن" else message.text.strip()
    await state.update_data(attr3=val)
    await state.set_state(AddProduct.attr4)
    await message.answer("خصوصیت ۴:", reply_markup=skip_kb())


@router.message(AddProduct.attr4)
async def add_attr4(message: Message, state: FSMContext):
    val = None if message.text == "⏭ رد کردن" else message.text.strip()
    await state.update_data(attr4=val)
    await state.set_state(AddProduct.attr5)
    await message.answer("خصوصیت ۵:", reply_markup=skip_kb())


@router.message(AddProduct.attr5)
async def add_attr5(message: Message, state: FSMContext):
    val = None if message.text == "⏭ رد کردن" else message.text.strip()
    await state.update_data(attr5=val)
    await state.set_state(AddProduct.brand)
    await message.answer("برند:", reply_markup=skip_kb())


@router.message(AddProduct.brand)
async def add_brand(message: Message, state: FSMContext):
    val = None if message.text == "⏭ رد کردن" else message.text.strip()
    await state.update_data(brand=val)
    await state.set_state(AddProduct.purchase_price)
    await message.answer("قیمت خرید (تومان) - عدد وارد کنید یا رد کنید:", reply_markup=skip_kb())


@router.message(AddProduct.purchase_price)
async def add_purchase(message: Message, state: FSMContext):
    val = parse_number(message.text)
    if val is None:
        await message.answer("عدد معتبر وارد کنید (فارسی یا انگلیسی) یا رد کنید.")
        return
    await state.update_data(purchase_price=val)
    await state.set_state(AddProduct.sale_price)
    await message.answer(
        "درصد سود فروش را وارد کنید (مثلاً ۳۰ یعنی ۳۰٪ اضافه به قیمت خرید):\n"
        "یا «رد کردن» برای صفر.",
        reply_markup=skip_kb()
    )


@router.message(AddProduct.sale_price)
async def add_sale(message: Message, state: FSMContext):
    data = await state.get_data()
    purchase = data.get("purchase_price") or 0
    if message.text == "⏭ رد کردن":
        percent = None
        sale = 0
    else:
        percent = parse_number(message.text)
        if percent is None:
            await message.answer("عدد معتبر (درصد) وارد کنید یا رد کنید.")
            return
        sale = purchase * (1 + percent / 100.0) if purchase else 0
    await state.update_data(sale_price=sale, sale_percent=percent)
    await state.set_state(AddProduct.wholesale_price)
    await message.answer(
        "درصد سود عمده را وارد کنید (مثلاً ۱۵):\nیا «رد کردن» برای صفر.",
        reply_markup=skip_kb()
    )


@router.message(AddProduct.wholesale_price)
async def add_wholesale(message: Message, state: FSMContext):
    data = await state.get_data()
    purchase = data.get("purchase_price") or 0
    if message.text == "⏭ رد کردن":
        percent = None
        wholesale = 0
    else:
        percent = parse_number(message.text)
        if percent is None:
            await message.answer("عدد معتبر (درصد) وارد کنید یا رد کنید.")
            return
        wholesale = purchase * (1 + percent / 100.0) if purchase else 0
    await state.update_data(wholesale_price=wholesale, wholesale_percent=percent)
    await state.set_state(AddProduct.stock)
    await message.answer("موجودی (عدد صحیح):", reply_markup=skip_kb())


@router.message(AddProduct.stock)
async def add_stock(message: Message, state: FSMContext):
    if message.text == "⏭ رد کردن":
        val = 0
    else:
        val = parse_int(message.text)
        if val is None:
            await message.answer("عدد صحیح وارد کنید یا رد کنید.")
            return
    await state.update_data(stock=val)
    await state.set_state(AddProduct.barcode)
    await message.answer(
        "بارکد کالا را تایپ کنید، یا با دکمه زیر اسکن کنید، یا رد کنید:",
        reply_markup=barcode_kb()
    )


@router.message(AddProduct.barcode, F.content_type == ContentType.WEB_APP_DATA)
async def add_barcode_scan(message: Message, state: FSMContext):
    barcode = (message.web_app_data.data or "").strip()
    await state.update_data(barcode=barcode or None)
    await state.set_state(AddProduct.seller_name)
    await message.answer(
        f"بارکد اسکن‌شده: <code>{barcode or '—'}</code>\nنام فروشنده:",
        parse_mode="HTML",
        reply_markup=skip_kb()
    )


@router.message(AddProduct.barcode)
async def add_barcode(message: Message, state: FSMContext):
    val = None if message.text == "⏭ رد کردن" else message.text.strip()
    await state.update_data(barcode=val)
    await state.set_state(AddProduct.seller_name)
    await message.answer("نام فروشنده:", reply_markup=skip_kb())


@router.message(AddProduct.seller_name)
async def add_seller_name(message: Message, state: FSMContext):
    val = None if message.text == "⏭ رد کردن" else message.text.strip()
    await state.update_data(seller_name=val)
    await state.set_state(AddProduct.seller_code)
    await message.answer("کد فروشنده:", reply_markup=skip_kb())


@router.message(AddProduct.seller_code)
async def add_seller_code(message: Message, state: FSMContext):
    val = None if message.text == "⏭ رد کردن" else message.text.strip()
    await state.update_data(seller_code=val)
    await state.set_state(AddProduct.purchase_day)
    await message.answer("روز تاریخ خرید (۱ تا ۳۱):", reply_markup=skip_kb())


@router.message(AddProduct.purchase_day)
async def add_purchase_day(message: Message, state: FSMContext):
    if message.text == "⏭ رد کردن":
        await state.update_data(purchase_day=None, purchase_month=None, purchase_year=None)
        await state.set_state(AddProduct.photo)
        await message.answer("عکس کالا را ارسال کنید (یا رد کردن):", reply_markup=skip_kb())
        return
    val = parse_int(message.text)
    if val is None or val < 1 or val > 31:
        await message.answer("روز معتبر (۱–۳۱) وارد کنید یا رد کنید.")
        return
    await state.update_data(purchase_day=val)
    await state.set_state(AddProduct.purchase_month)
    await message.answer("ماه تاریخ خرید (۱ تا ۱۲):", reply_markup=skip_kb())


@router.message(AddProduct.purchase_month)
async def add_purchase_month(message: Message, state: FSMContext):
    if message.text == "⏭ رد کردن":
        await state.update_data(purchase_month=None, purchase_year=None)
        await state.set_state(AddProduct.photo)
        await message.answer("عکس کالا را ارسال کنید (یا رد کردن):", reply_markup=skip_kb())
        return
    val = parse_int(message.text)
    if val is None or val < 1 or val > 12:
        await message.answer("ماه معتبر (۱–۱۲) وارد کنید یا رد کنید.")
        return
    await state.update_data(purchase_month=val)
    await state.set_state(AddProduct.purchase_year)
    await message.answer(
        f"سال تاریخ خرید (هجری شمسی).\nمی‌توانید {CURRENT_YEAR} را انتخاب کنید یا سال دیگری تایپ کنید:",
        reply_markup=year_suggest_kb(CURRENT_YEAR)
    )


@router.message(AddProduct.purchase_year)
async def add_purchase_year(message: Message, state: FSMContext):
    if message.text == "⏭ رد کردن":
        await state.update_data(purchase_year=None)
    else:
        val = parse_int(message.text)
        if val is None or val < 1300 or val > 1500:
            await message.answer("سال معتبر وارد کنید (مثلاً ۱۴۰۵) یا رد کنید.")
            return
        await state.update_data(purchase_year=val)
    await state.set_state(AddProduct.photo)
    await message.answer("عکس کالا را ارسال کنید (یا رد کردن):", reply_markup=skip_kb())


@router.message(AddProduct.photo, F.photo)
async def add_photo(message: Message, state: FSMContext):
    file_id = message.photo[-1].file_id
    data = await state.get_data()
    data["photo_file_id"] = file_id
    product_id = await add_product(data)
    await state.clear()
    product = await get_product(product_id)
    await message.answer("✅ کالا با موفقیت اضافه شد:", reply_markup=admin_menu())
    await message.answer(format_product(product), parse_mode="HTML",
                         reply_markup=product_actions_kb(product_id, True))


@router.message(AddProduct.photo)
async def add_photo_skip(message: Message, state: FSMContext):
    if message.text != "⏭ رد کردن":
        await message.answer("عکس بفرستید یا رد کنید.")
        return
    data = await state.get_data()
    data["photo_file_id"] = None
    product_id = await add_product(data)
    await state.clear()
    product = await get_product(product_id)
    await message.answer("✅ کالا با موفقیت اضافه شد (بدون عکس):", reply_markup=admin_menu())
    await message.answer(format_product(product), parse_mode="HTML",
                         reply_markup=product_actions_kb(product_id, True))


# ---------- ویرایش کالا ----------
FIELD_FA_NAMES = {
    "category": "دسته",
    "attr1": "خصوصیت ۱",
    "attr2": "خصوصیت ۲",
    "attr3": "خصوصیت ۳",
    "attr4": "خصوصیت ۴",
    "attr5": "خصوصیت ۵",
    "brand": "برند",
    "purchase_price": "قیمت خرید",
    "sale_price": "درصد سود فروش",
    "wholesale_price": "درصد سود عمده",
    "stock": "موجودی",
    "barcode": "بارکد",
    "seller_name": "فروشنده",
    "seller_code": "کد فروشنده",
    "photo_file_id": "عکس",
    "purchase_day": "روز خرید",
    "purchase_month": "ماه خرید",
    "purchase_year": "سال خرید",
}


@router.message(F.text == "✏️ ویرایش کالا")
@admin_only
async def edit_start(message: Message, state: FSMContext):
    await state.set_state(EditProduct.waiting_id)
    await message.answer(
        "شناسه کالایی که می‌خواهید ویرایش کنید را وارد کنید:\n"
        "(همان شناسه‌ای که زیر مشخصات هر کالا نمایش داده می‌شود)",
        reply_markup=cancel_kb()
    )


@router.message(EditProduct.waiting_id, F.text == "❌ انصراف")
async def edit_cancel(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("لغو شد.", reply_markup=main_menu(is_admin(message.from_user)))


@router.message(EditProduct.waiting_id)
async def edit_id(message: Message, state: FSMContext):
    excel_id = normalize_digits(message.text.strip())
    if not excel_id:
        await message.answer("شناسه را وارد کنید.")
        return
    product = await get_product_by_excel_id(excel_id)
    if not product:
        await message.answer("کالایی با این شناسه پیدا نشد.")
        return
    pid = product["id"]
    await state.clear()
    await message.answer(format_product(product), parse_mode="HTML")
    await message.answer(
        "کدام فیلد را می‌خواهید تغییر دهید؟",
        reply_markup=edit_fields_kb(pid)
    )


@router.callback_query(F.data.startswith("edit:"))
@admin_only
async def inline_edit_start(callback: CallbackQuery, state: FSMContext):
    pid = int(callback.data.split(":")[1])
    product = await get_product(pid)
    if not product:
        await callback.answer("کالا پیدا نشد", show_alert=True)
        return
    await state.clear()
    await callback.message.answer(format_product(product), parse_mode="HTML")
    await callback.message.answer(
        "کدام فیلد را می‌خواهید تغییر دهید؟",
        reply_markup=edit_fields_kb(pid)
    )
    await callback.answer()


@router.callback_query(F.data == "cancel_edit")
async def cancel_edit_cb(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.edit_text("ویرایش لغو شد.")
    await callback.message.answer("منوی اصلی:", reply_markup=main_menu(is_admin(callback.from_user)))
    await callback.answer()


@router.callback_query(F.data.startswith("efield:"))
@admin_only
async def edit_field_chosen(callback: CallbackQuery, state: FSMContext):
    try:
        _, pid_str, field = callback.data.split(":", 2)
        pid = int(pid_str)
    except Exception:
        await callback.answer("خطا", show_alert=True)
        return

    await state.update_data(edit_id=pid, edit_field=field)
    fa_name = FIELD_FA_NAMES.get(field, field)

    if field == "photo_file_id":
        await callback.message.answer(
            f"🖼 عکس جدید برای «{fa_name}» را ارسال کنید:",
            reply_markup=cancel_kb()
        )
    elif field == "barcode":
        await callback.message.answer(
            "بارکد جدید را تایپ کنید یا با دکمه زیر اسکن کنید:",
            reply_markup=barcode_kb()
        )
    elif field == "sale_price":
        await callback.message.answer(
            "درصد سود فروش جدید را وارد کنید (مثلاً ۳۰):",
            reply_markup=cancel_kb()
        )
    elif field == "wholesale_price":
        await callback.message.answer(
            "درصد سود عمده جدید را وارد کنید (مثلاً ۱۵):",
            reply_markup=cancel_kb()
        )
    elif field == "purchase_year":
        await callback.message.answer(
            f"سال خرید جدید را وارد کنید (پیشنهاد: {CURRENT_YEAR}):",
            reply_markup=year_suggest_kb(CURRENT_YEAR)
        )
    else:
        await callback.message.answer(
            f"مقدار جدید برای «{fa_name}» را وارد کنید:",
            reply_markup=cancel_kb()
        )
    await state.set_state(EditProduct.value)
    await callback.answer()


@router.message(EditProduct.value, F.text == "❌ انصراف")
async def edit_value_cancel(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("لغو شد.", reply_markup=main_menu(is_admin(message.from_user)))


@router.message(EditProduct.value, F.content_type == ContentType.WEB_APP_DATA)
async def edit_barcode_scan(message: Message, state: FSMContext):
    data = await state.get_data()
    if data.get("edit_field") != "barcode":
        await message.answer("الان منتظر ورودی دیگری هستم.")
        return
    pid = data["edit_id"]
    product = await get_product(pid)
    if not product:
        await state.clear()
        await message.answer("کالا پیدا نشد.", reply_markup=admin_menu())
        return
    barcode = (message.web_app_data.data or "").strip()
    product["barcode"] = barcode or None
    await update_product(pid, product)
    await state.clear()
    updated = await get_product(pid)
    await message.answer(
        f"✅ بارکد به‌روز شد: <code>{barcode or '—'}</code>",
        parse_mode="HTML",
        reply_markup=admin_menu()
    )
    await message.answer(
        format_product(updated),
        parse_mode="HTML",
        reply_markup=product_actions_kb(pid, True)
    )


@router.message(EditProduct.value, F.photo)
async def edit_photo(message: Message, state: FSMContext):
    data = await state.get_data()
    if data.get("edit_field") != "photo_file_id":
        await message.answer("الان منتظر متن هستم.")
        return
    pid = data["edit_id"]
    product = await get_product(pid)
    if not product:
        await state.clear()
        await message.answer("کالا پیدا نشد.", reply_markup=admin_menu())
        return
    product["photo_file_id"] = message.photo[-1].file_id
    await update_product(pid, product)
    await state.clear()
    await message.answer("✅ عکس به‌روز شد.", reply_markup=admin_menu())
    await message.answer_photo(product["photo_file_id"], caption=format_product(product), parse_mode="HTML")


@router.message(EditProduct.value)
async def edit_value(message: Message, state: FSMContext):
    data = await state.get_data()
    field = data.get("edit_field")
    pid = data["edit_id"]
    product = await get_product(pid)
    if not product:
        await state.clear()
        await message.answer("کالا پیدا نشد.", reply_markup=admin_menu())
        return

    text = message.text.strip()
    purchase = product.get("purchase_price") or 0

    if field == "purchase_price":
        val = parse_number(text)
        if val is None:
            await message.answer("عدد معتبر وارد کنید.")
            return
        product["purchase_price"] = val
        # اگر درصدها موجودند، فروش و عمده را دوباره محاسبه کن
        if product.get("sale_percent") is not None:
            product["sale_price"] = val * (1 + float(product["sale_percent"]) / 100.0)
        if product.get("wholesale_percent") is not None:
            product["wholesale_price"] = val * (1 + float(product["wholesale_percent"]) / 100.0)

    elif field == "sale_price":
        percent = parse_number(text)
        if percent is None:
            await message.answer("درصد معتبر وارد کنید.")
            return
        product["sale_percent"] = percent
        product["sale_price"] = purchase * (1 + percent / 100.0) if purchase else 0

    elif field == "wholesale_price":
        percent = parse_number(text)
        if percent is None:
            await message.answer("درصد معتبر وارد کنید.")
            return
        product["wholesale_percent"] = percent
        product["wholesale_price"] = purchase * (1 + percent / 100.0) if purchase else 0

    elif field == "stock":
        val = parse_int(text)
        if val is None:
            await message.answer("عدد صحیح وارد کنید.")
            return
        product[field] = val

    elif field in ("purchase_day", "purchase_month", "purchase_year"):
        val = parse_int(text)
        if val is None:
            await message.answer("عدد معتبر وارد کنید.")
            return
        if field == "purchase_day" and (val < 1 or val > 31):
            await message.answer("روز باید بین ۱ تا ۳۱ باشد.")
            return
        if field == "purchase_month" and (val < 1 or val > 12):
            await message.answer("ماه باید بین ۱ تا ۱۲ باشد.")
            return
        product[field] = val

    elif field == "barcode":
        if text == "⏭ رد کردن":
            product["barcode"] = None
        else:
            product["barcode"] = text
    else:
        product[field] = text

    await update_product(pid, product)
    await state.clear()
    updated = await get_product(pid)
    await message.answer("✅ کالا به‌روز شد:", reply_markup=admin_menu())
    await message.answer(format_product(updated), parse_mode="HTML",
                         reply_markup=product_actions_kb(pid, True))


# ---------- هیستوری قیمت ----------
@router.callback_query(F.data.startswith("addhist:"))
@admin_only
async def add_history_start(callback: CallbackQuery, state: FSMContext):
    pid = int(callback.data.split(":")[1])
    product = await get_product(pid)
    if not product:
        await callback.answer("کالا پیدا نشد", show_alert=True)
        return
    await state.update_data(hist_product_id=pid)
    await state.set_state(AddHistory.day)
    await callback.message.answer(
        f"افزودن هیستوری قیمت خرید برای:\n{product.get('name') or '—'}\n\n"
        "روز تاریخ را وارد کنید (۱–۳۱):",
        reply_markup=cancel_kb()
    )
    await callback.answer()


@router.message(AddHistory.day)
async def hist_day(message: Message, state: FSMContext):
    val = parse_int(message.text)
    if val is None or val < 1 or val > 31:
        await message.answer("روز معتبر (۱–۳۱) وارد کنید.")
        return
    await state.update_data(hist_day=val)
    await state.set_state(AddHistory.month)
    await message.answer("ماه را وارد کنید (۱–۱۲):", reply_markup=cancel_kb())


@router.message(AddHistory.month)
async def hist_month(message: Message, state: FSMContext):
    val = parse_int(message.text)
    if val is None or val < 1 or val > 12:
        await message.answer("ماه معتبر (۱–۱۲) وارد کنید.")
        return
    await state.update_data(hist_month=val)
    await state.set_state(AddHistory.year)
    await message.answer(
        f"سال را وارد کنید (پیشنهاد: {CURRENT_YEAR}):",
        reply_markup=year_suggest_kb(CURRENT_YEAR)
    )


@router.message(AddHistory.year)
async def hist_year(message: Message, state: FSMContext):
    val = parse_int(message.text)
    if val is None or val < 1300 or val > 1500:
        await message.answer("سال معتبر وارد کنید.")
        return
    await state.update_data(hist_year=val)
    await state.set_state(AddHistory.price)
    await message.answer("قیمت خرید جدید را وارد کنید (تومان):", reply_markup=cancel_kb())


@router.message(AddHistory.price)
async def hist_price(message: Message, state: FSMContext):
    price = parse_number(message.text)
    if price is None or price < 0:
        await message.answer("قیمت معتبر وارد کنید.")
        return
    data = await state.get_data()
    pid = data["hist_product_id"]
    day = data["hist_day"]
    month = data["hist_month"]
    year = data["hist_year"]

    await add_price_history(pid, day, month, year, price)
    await state.clear()

    product = await get_product(pid)
    await message.answer(
        f"✅ هیستوری ثبت شد.\n"
        f"📅 {day}/{month}/{year} — {format_price(price)}\n\n"
        "اگر تاریخ جدیدتر بود، قیمت خرید و فروش/عمده محصول به‌روز شدند.",
        reply_markup=admin_menu()
    )
    await message.answer(
        format_product(product),
        parse_mode="HTML",
        reply_markup=product_actions_kb(pid, True)
    )


@router.callback_query(F.data.startswith("showhist:"))
async def show_history(callback: CallbackQuery):
    """نمایش هیستوری برای همه؛ دکمه‌های ویرایش/حذف فقط برای ادمین"""
    pid = int(callback.data.split(":")[1])
    product = await get_product(pid)
    if not product:
        await callback.answer("کالا پیدا نشد", show_alert=True)
        return

    history = await get_price_history(pid)
    if not history:
        await callback.message.answer(
            f"📜 هیستوری‌ای برای «{product.get('name') or '—'}» ثبت نشده."
        )
        await callback.answer()
        return

    admin = is_admin(callback.from_user)
    await callback.message.answer(
        f"📜 هیستوری قیمت خرید — <b>{product.get('name') or '—'}</b>",
        parse_mode="HTML"
    )
    for h in history:
        line = f"📅 {h['day']}/{h['month']}/{h['year']} — {format_price(h['price'])}"
        kb = history_item_kb(h["id"], pid, admin)
        await callback.message.answer(line, reply_markup=kb)
    await callback.answer()


# ---------- ویرایش هیستوری ----------
@router.callback_query(F.data.startswith("edithist:"))
@admin_only
async def edit_history_start(callback: CallbackQuery, state: FSMContext):
    hid = int(callback.data.split(":")[1])
    item = await get_history_item(hid)
    if not item:
        await callback.answer("رکورد پیدا نشد", show_alert=True)
        return
    await state.update_data(
        edit_hist_id=hid,
        edit_hist_product_id=item["product_id"],
    )
    await state.set_state(EditHistory.day)
    await callback.message.answer(
        f"ویرایش هیستوری فعلی: {item['day']}/{item['month']}/{item['year']} — {format_price(item['price'])}\n\n"
        "روز جدید را وارد کنید (۱–۳۱):",
        reply_markup=cancel_kb()
    )
    await callback.answer()


@router.message(EditHistory.day, F.text == "❌ انصراف")
@router.message(EditHistory.month, F.text == "❌ انصراف")
@router.message(EditHistory.year, F.text == "❌ انصراف")
@router.message(EditHistory.price, F.text == "❌ انصراف")
async def edit_hist_cancel(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("لغو شد.", reply_markup=main_menu(is_admin(message.from_user)))


@router.message(EditHistory.day)
async def edit_hist_day(message: Message, state: FSMContext):
    val = parse_int(message.text)
    if val is None or val < 1 or val > 31:
        await message.answer("روز معتبر (۱–۳۱) وارد کنید.")
        return
    await state.update_data(edit_hist_day=val)
    await state.set_state(EditHistory.month)
    await message.answer("ماه جدید (۱–۱۲):", reply_markup=cancel_kb())


@router.message(EditHistory.month)
async def edit_hist_month(message: Message, state: FSMContext):
    val = parse_int(message.text)
    if val is None or val < 1 or val > 12:
        await message.answer("ماه معتبر (۱–۱۲) وارد کنید.")
        return
    await state.update_data(edit_hist_month=val)
    await state.set_state(EditHistory.year)
    await message.answer(
        f"سال جدید (پیشنهاد: {CURRENT_YEAR}):",
        reply_markup=year_suggest_kb(CURRENT_YEAR)
    )


@router.message(EditHistory.year)
async def edit_hist_year(message: Message, state: FSMContext):
    val = parse_int(message.text)
    if val is None or val < 1300 or val > 1500:
        await message.answer("سال معتبر وارد کنید.")
        return
    await state.update_data(edit_hist_year=val)
    await state.set_state(EditHistory.price)
    await message.answer("قیمت خرید جدید (تومان):", reply_markup=cancel_kb())


@router.message(EditHistory.price)
async def edit_hist_price(message: Message, state: FSMContext):
    price = parse_number(message.text)
    if price is None or price < 0:
        await message.answer("قیمت معتبر وارد کنید.")
        return
    data = await state.get_data()
    hid = data["edit_hist_id"]
    pid = data["edit_hist_product_id"]
    day = data["edit_hist_day"]
    month = data["edit_hist_month"]
    year = data["edit_hist_year"]

    await update_price_history(hid, day, month, year, price)
    await sync_product_from_latest_history(pid)
    await state.clear()

    product = await get_product(pid)
    await message.answer(
        f"✅ هیستوری ویرایش شد: {day}/{month}/{year} — {format_price(price)}\n"
        "قیمت کالا بر اساس جدیدترین هیستوری به‌روز شد.",
        reply_markup=admin_menu()
    )
    if product:
        await message.answer(
            format_product(product),
            parse_mode="HTML",
            reply_markup=product_actions_kb(pid, True)
        )


# ---------- حذف هیستوری ----------
@router.callback_query(F.data.startswith("delhist:"))
@admin_only
async def delete_history_ask(callback: CallbackQuery):
    parts = callback.data.split(":")
    hid = int(parts[1])
    pid = int(parts[2])
    item = await get_history_item(hid)
    if not item:
        await callback.answer("رکورد پیدا نشد", show_alert=True)
        return
    await callback.message.answer(
        f"حذف این رکورد هیستوری؟\n📅 {item['day']}/{item['month']}/{item['year']} — {format_price(item['price'])}",
        reply_markup=confirm_delete_history_kb(hid, pid)
    )
    await callback.answer()


@router.callback_query(F.data.startswith("confirm_delhist:"))
@admin_only
async def confirm_delete_history(callback: CallbackQuery):
    parts = callback.data.split(":")
    hid = int(parts[1])
    pid = int(parts[2])
    await delete_price_history(hid)
    await sync_product_from_latest_history(pid)
    product = await get_product(pid)
    await callback.message.edit_text("✅ رکورد هیستوری حذف شد. قیمت کالا در صورت نیاز به‌روز شد.")
    if product:
        await callback.message.answer(
            format_product(product),
            parse_mode="HTML",
            reply_markup=product_actions_kb(pid, True)
        )
    await callback.answer()


@router.callback_query(F.data == "cancel_delhist")
async def cancel_delete_history(callback: CallbackQuery):
    await callback.message.edit_text("حذف هیستوری لغو شد.")
    await callback.answer()


# ---------- جستجو بر اساس شناسه ----------
@router.message(F.text == "🔎 جستجو بر اساس شناسه")
@admin_only
async def id_search_start(message: Message, state: FSMContext):
    await state.set_state(IdSearch.waiting_id)
    await message.answer(
        "شناسه کالا (ستون شناسه اکسل) را وارد کنید:\n"
        "ارقام فارسی و انگلیسی هر دو پذیرفته می‌شوند.",
        reply_markup=cancel_kb()
    )


@router.message(IdSearch.waiting_id, F.text == "❌ انصراف")
async def id_search_cancel(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("لغو شد.", reply_markup=main_menu(is_admin(message.from_user)))


@router.message(IdSearch.waiting_id)
async def id_search_query(message: Message, state: FSMContext):
    excel_id = normalize_digits(message.text.strip())
    if not excel_id:
        await message.answer("شناسه را وارد کنید.")
        return
    product = await get_product_by_excel_id(excel_id)
    if not product:
        await message.answer(
            f"❌ کالایی با شناسه <code>{excel_id}</code> پیدا نشد.",
            parse_mode="HTML",
            reply_markup=cancel_kb()
        )
        return
    await state.clear()
    admin = is_admin(message.from_user)
    text = format_product(product)
    kb = product_actions_kb(product["id"], admin)
    if product.get("photo_file_id"):
        await message.answer_photo(product["photo_file_id"], caption=text, parse_mode="HTML", reply_markup=kb)
    else:
        await message.answer(text, parse_mode="HTML", reply_markup=kb)
    await message.answer("منوی مدیریت:", reply_markup=admin_menu())


# ---------- حذف کالا ----------
@router.message(F.text == "🗑 حذف کالا")
@admin_only
async def delete_start(message: Message, state: FSMContext):
    await state.set_state(DeleteProduct.waiting_id)
    await message.answer("شناسه کالایی که می‌خواهید حذف کنید را وارد کنید:", reply_markup=cancel_kb())


@router.message(DeleteProduct.waiting_id, F.text == "❌ انصراف")
async def delete_cancel(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("لغو شد.", reply_markup=main_menu(is_admin(message.from_user)))


@router.message(DeleteProduct.waiting_id)
async def delete_id(message: Message, state: FSMContext):
    excel_id = normalize_digits(message.text.strip())
    if not excel_id:
        await message.answer("شناسه را وارد کنید.")
        return
    product = await get_product_by_excel_id(excel_id)
    if not product:
        await message.answer("کالایی پیدا نشد.")
        return
    await state.clear()
    await message.answer(
        f"آیا مطمئن هستید که می‌خواهید این کالا را حذف کنید؟\n\n{format_product(product)}",
        parse_mode="HTML",
        reply_markup=confirm_delete_kb(product["id"])
    )


@router.callback_query(F.data.startswith("confirm_del:"))
@admin_only
async def confirm_delete(callback: CallbackQuery):
    pid = int(callback.data.split(":")[1])
    await delete_product(pid)
    await callback.message.edit_text("✅ کالا حذف شد.")
    await callback.answer()


@router.callback_query(F.data == "cancel_del")
async def cancel_delete(callback: CallbackQuery):
    await callback.message.edit_text("حذف لغو شد.")
    await callback.message.answer("منوی اصلی:", reply_markup=main_menu(is_admin(callback.from_user)))
    await callback.answer()


@router.callback_query(F.data.startswith("del:"))
@admin_only
async def inline_delete(callback: CallbackQuery):
    pid = int(callback.data.split(":")[1])
    product = await get_product(pid)
    if not product:
        await callback.answer("پیدا نشد", show_alert=True)
        return
    await callback.message.answer(
        f"حذف این کالا؟\n\n{format_product(product)}",
        parse_mode="HTML",
        reply_markup=confirm_delete_kb(pid)
    )
    await callback.answer()


# ---------- لیست کالاها ----------
@router.message(F.text == "📋 لیست کالاها")
@admin_only
async def list_products_menu(message: Message, state: FSMContext):
    await state.clear()
    await message.answer(
        "چطور می‌خواهید کالاها را ببینید؟",
        reply_markup=list_products_menu_kb()
    )


@router.callback_query(F.data == "browse_by_category")
@admin_only
async def browse_by_category_start(callback: CallbackQuery, state: FSMContext):
    categories = await get_distinct_categories()
    if not categories:
        await callback.message.edit_text("هنوز کالایی ثبت نشده.")
        await callback.answer()
        return
    await state.update_data(browse_categories=categories)
    await callback.message.edit_text(
        f"📋 {len(categories)} دسته موجود است. یک دسته را انتخاب کنید:",
        reply_markup=browse_categories_kb(categories, page=0)
    )
    await callback.answer()


@router.callback_query(F.data == "browse_by_seller")
@admin_only
async def browse_by_seller_start(callback: CallbackQuery, state: FSMContext):
    await state.set_state(SellerSearch.waiting_name)
    await callback.message.answer("نام فروشنده را وارد کنید:", reply_markup=cancel_kb())
    await callback.answer()


@router.message(SellerSearch.waiting_name, F.text == "❌ انصراف")
async def seller_search_cancel(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("لغو شد.", reply_markup=main_menu(is_admin(message.from_user)))


@router.message(SellerSearch.waiting_name)
async def seller_search_query(message: Message, state: FSMContext):
    query = message.text.strip()
    if not query:
        await message.answer("لطفاً نام فروشنده را وارد کنید.")
        return

    total = await count_by_seller(query)
    if total == 0:
        await message.answer(
            "❌ کالایی از این فروشنده پیدا نشد.\nنام دیگری امتحان کنید یا انصراف بزنید.",
            reply_markup=cancel_kb()
        )
        return

    await state.update_data(seller_query=query)
    products = await search_by_seller(query, offset=0, limit=PAGE_SIZE)
    await message.answer(
        f"👤 کالاهای فروشنده «{query}» ({total} مورد):",
        reply_markup=results_kb(products, 0, total, PAGE_SIZE, "spage")
    )


@router.callback_query(F.data.startswith("spage:"))
@admin_only
async def seller_pagination_handler(callback: CallbackQuery, state: FSMContext):
    try:
        page = int(callback.data.split(":", 1)[1])
    except Exception:
        await callback.answer("خطا در صفحه‌بندی")
        return

    data = await state.get_data()
    query = data.get("seller_query")
    if not query:
        await callback.answer("عبارت جستجو منقضی شده، دوباره جستجو کنید.", show_alert=True)
        return

    total = await count_by_seller(query)
    products = await search_by_seller(query, offset=page * PAGE_SIZE, limit=PAGE_SIZE)

    await callback.message.edit_text(
        f"👤 کالاهای فروشنده «{query}» ({total} مورد) - صفحه {page + 1}:",
        reply_markup=results_kb(products, page, total, PAGE_SIZE, "spage")
    )
    await callback.answer()


@router.callback_query(F.data.startswith("browsecatpage:"))
@admin_only
async def browse_category_page(callback: CallbackQuery, state: FSMContext):
    page = int(callback.data.split(":", 1)[1])
    data = await state.get_data()
    categories = data.get("browse_categories") or []
    await callback.message.edit_reply_markup(reply_markup=browse_categories_kb(categories, page=page))
    await callback.answer()


@router.callback_query(F.data.startswith("browsecat:"))
@admin_only
async def browse_category_select(callback: CallbackQuery, state: FSMContext):
    idx = int(callback.data.split(":", 1)[1])
    data = await state.get_data()
    categories = data.get("browse_categories") or []
    if idx < 0 or idx >= len(categories):
        await callback.answer("این لیست منقضی شده؛ دوباره روی «📋 لیست کالاها» بزنید.", show_alert=True)
        return

    category = categories[idx]
    attrs = await get_distinct_attr1(category)
    await state.update_data(browse_category=category, browse_attrs=attrs)

    if not attrs:
        await state.update_data(browse_attr1=None)
        await callback.answer()
        await send_browse_products(callback.message, state, callback.from_user, page=0)
        return

    await callback.message.edit_text(
        f"🏷 دسته: {category}\nخصوصیت ۱ را انتخاب کنید:",
        reply_markup=browse_attrs_kb(attrs, page=0)
    )
    await callback.answer()


@router.callback_query(F.data.startswith("browseattrpage:"))
@admin_only
async def browse_attr_page(callback: CallbackQuery, state: FSMContext):
    page = int(callback.data.split(":", 1)[1])
    data = await state.get_data()
    attrs = data.get("browse_attrs") or []
    await callback.message.edit_reply_markup(reply_markup=browse_attrs_kb(attrs, page=page))
    await callback.answer()


@router.callback_query(F.data.startswith("browseattr:"))
@admin_only
async def browse_attr_select(callback: CallbackQuery, state: FSMContext):
    idx = int(callback.data.split(":", 1)[1])
    data = await state.get_data()
    attrs = data.get("browse_attrs") or []
    category = data.get("browse_category")
    if not category or idx < 0 or idx >= len(attrs):
        await callback.answer("این لیست منقضی شده؛ دوباره روی «📋 لیست کالاها» بزنید.", show_alert=True)
        return

    attr1 = attrs[idx]
    await state.update_data(browse_attr1=attr1)
    await callback.answer()
    await send_browse_products(callback.message, state, callback.from_user, page=0)


@router.callback_query(F.data == "browseback:cat")
@admin_only
async def browse_back_to_categories(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    categories = data.get("browse_categories")
    if not categories:
        categories = await get_distinct_categories()
        await state.update_data(browse_categories=categories)
    await callback.message.edit_text(
        f"📋 {len(categories)} دسته موجود است. یک دسته را انتخاب کنید:",
        reply_markup=browse_categories_kb(categories, page=0)
    )
    await callback.answer()


@router.callback_query(F.data == "browseback:attr")
@admin_only
async def browse_back_to_attrs(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    category = data.get("browse_category") or ""
    attrs = data.get("browse_attrs") or []
    await callback.message.edit_text(
        f"🏷 دسته: {category}\nخصوصیت ۱ را انتخاب کنید:",
        reply_markup=browse_attrs_kb(attrs, page=0)
    )
    await callback.answer()


@router.callback_query(F.data.startswith("browsepage:"))
@admin_only
async def browse_products_page(callback: CallbackQuery, state: FSMContext):
    page = int(callback.data.split(":", 1)[1])
    await callback.answer()
    await send_browse_products(callback.message, state, callback.from_user, page=page)


async def send_browse_products(message: Message, state: FSMContext, user, page: int):
    data = await state.get_data()
    category = data.get("browse_category")
    attr1 = data.get("browse_attr1")
    has_attr = "browse_attrs" in data and bool(data.get("browse_attrs"))

    if not category:
        await message.answer("دسته انتخاب نشده؛ دوباره روی «📋 لیست کالاها» بزنید.")
        return

    total = await count_products_by_category_attr1(category, attr1)
    products = await get_products_by_category_attr1(category, attr1, offset=page * PAGE_SIZE, limit=PAGE_SIZE)
    admin = is_admin(user)

    title = f"🏷 {category}" + (f" | {attr1}" if attr1 else "")
    await message.answer(f"{title}\n{total} کالا (صفحه {page + 1}):")

    for p in products:
        text = format_product(p)
        kb = product_actions_kb(p["id"], admin)
        if p.get("photo_file_id"):
            await message.answer_photo(p["photo_file_id"], caption=text, parse_mode="HTML", reply_markup=kb)
        else:
            await message.answer(text, parse_mode="HTML", reply_markup=kb)

    await message.answer(
        "ناوبری:",
        reply_markup=browse_products_nav_kb(page, total, PAGE_SIZE, has_attr=has_attr and attr1 is not None)
    )


# ---------- پاک کردن کامل دیتابیس ----------
@router.message(F.text == "🧨 پاک کردن کامل دیتابیس")
@admin_only
async def wipe_start(message: Message):
    total = await count_products()
    await message.answer(
        f"⚠️ این عملیات همه‌ی {total} کالای ثبت‌شده را برای همیشه پاک می‌کند و "
        "غیرقابل بازگشت است.\nمطمئن هستید؟",
        reply_markup=confirm_wipe_kb()
    )


@router.callback_query(F.data == "confirm_wipe")
@admin_only
async def confirm_wipe(callback: CallbackQuery):
    await clear_products()
    await callback.message.edit_text("✅ کل دیتابیس پاک شد. می‌توانید از صفر کالا اضافه کنید.")
    await callback.message.answer("منوی مدیریت:", reply_markup=admin_menu())
    await callback.answer()


@router.callback_query(F.data == "cancel_wipe")
async def cancel_wipe(callback: CallbackQuery):
    await callback.message.edit_text("لغو شد؛ چیزی پاک نشد.")
    await callback.message.answer("منوی مدیریت:", reply_markup=admin_menu())
    await callback.answer()


# ---------- خروجی اکسل ----------
@router.message(F.text == "📤 خروجی اکسل")
@admin_only
async def export_excel(message: Message):
    products = await get_all_products_full()
    if not products:
        await message.answer("هنوز کالایی ثبت نشده.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "کالاها"
    headers = [
        "شناسه", "دسته", "خصوصیت 1", "خصوصیت 2", "خصوصیت 3", "خصوصیت 4", "خصوصیت 5",
        "برند", "قیمت خرید", "قیمت فروش واحد تک", "قیمت فروش عمده", "بارکد",
        "موجودی", "فروشنده", "کد فروشنده",
        "تاریخ خرید - روز", "تاریخ خرید - ماه", "تاریخ خرید - سال",
    ]
    ws.append(headers)
    for p in products:
        ws.append([
            p.get("excel_id") or "",
            p.get("category") or "",
            p.get("attr1") or "",
            p.get("attr2") or "",
            p.get("attr3") or "",
            p.get("attr4") or "",
            p.get("attr5") or "",
            p.get("brand") or "",
            p.get("purchase_price") or 0,
            p.get("sale_price") or 0,
            p.get("wholesale_price") or 0,
            p.get("barcode") or "",
            p.get("stock") or 0,
            p.get("seller_name") or "",
            p.get("seller_code") or "",
            p.get("purchase_day") or "",
            p.get("purchase_month") or "",
            p.get("purchase_year") or "",
        ])

    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name
        await message.answer_document(
            FSInputFile(tmp_path, filename="kalaha.xlsx"),
            caption=f"📤 خروجی {len(products)} کالا (شامل تاریخ خرید و بارکد)."
        )
    finally:
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except Exception:
                pass


# ---------- ایمپورت اکسل ----------
@router.message(F.text == "📥 ایمپورت اکسل")
@admin_only
async def import_start(message: Message, state: FSMContext):
    await state.set_state(ImportExcel.waiting_file)
    await message.answer(
        "فایل اکسل را ارسال کنید.\n"
        "✅ کالاهای موجود فقط به‌روزرسانی می‌شوند و بارکد/موجودی/عکس قبلی حفظ می‌شود.\n"
        "🆔 تشخیص تکراری بر اساس ستون «شناسه» است.\n"
        "ستون‌های تاریخ خرید (روز/ماه/سال) و بارکد پشتیبانی می‌شوند.",
        reply_markup=cancel_kb()
    )


@router.message(ImportExcel.waiting_file, F.text == "❌ انصراف")
async def import_cancel(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("لغو شد.", reply_markup=main_menu(is_admin(message.from_user)))


@router.message(ImportExcel.waiting_file, F.document)
async def import_excel(message: Message, state: FSMContext):
    doc = message.document
    if not doc.file_name.endswith((".xlsx", ".xls")):
        await message.answer("فقط فایل اکسل (.xlsx) بفرستید.")
        return

    await message.answer("در حال پردازش فایل... لطفاً صبر کنید.")
    file = await message.bot.get_file(doc.file_id)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        await message.bot.download_file(file.file_path, tmp.name)
        tmp_path = tmp.name

    try:
        wb = openpyxl.load_workbook(tmp_path)
        ws = wb.active

        header_row = None
        for row in range(1, 6):
            vals = [ws.cell(row, c).value for c in range(1, 20)]
            if any(v and "دسته" in str(v) for v in vals):
                header_row = row
                break
        if not header_row:
            await message.answer("هدر «دسته» در فایل پیدا نشد.")
            return

        col_map = {}
        for c in range(1, 25):
            val = ws.cell(header_row, c).value
            if not val:
                continue
            v = str(val).strip()
            if "شناسه" in v:
                col_map["excel_id"] = c
            elif "دسته" in v:
                col_map["category"] = c
            elif "خصوصیت 1" in v or "خصوصیت۱" in v:
                col_map["attr1"] = c
            elif "خصوصیت 2" in v or "خصوصیت۲" in v:
                col_map["attr2"] = c
            elif "خصوصیت 3" in v or "خصوصیت۳" in v:
                col_map["attr3"] = c
            elif "خصوصیت 4" in v or "خصوصیت۴" in v:
                col_map["attr4"] = c
            elif "خصوصیت 5" in v or "خصوصیت۵" in v:
                col_map["attr5"] = c
            elif "برند" in v:
                col_map["brand"] = c
            elif "قیمت خرید" in v:
                col_map["purchase_price"] = c
            elif "فروش واحد تک" in v or "قیمت فروش واحد تک" in v:
                col_map["sale_price"] = c
            elif "عمده" in v:
                col_map["wholesale_price"] = c
            elif "فروشنده" in v and "کد" not in v:
                col_map["seller_name"] = c
            elif "کد فروشنده" in v:
                col_map["seller_code"] = c
            elif "بارکد" in v:
                col_map["barcode"] = c
            elif "موجودی" in v:
                col_map["stock"] = c
            elif "تاریخ خرید - روز" in v or ("روز" in v and "تاریخ" in v):
                col_map["purchase_day"] = c
            elif "تاریخ خرید - ماه" in v or ("ماه" in v and "تاریخ" in v):
                col_map["purchase_month"] = c
            elif "تاریخ خرید - سال" in v or ("سال" in v and "تاریخ" in v):
                col_map["purchase_year"] = c

        if "category" not in col_map:
            await message.answer("ستون دسته پیدا نشد.")
            return

        added, updated = 0, 0
        for row in range(header_row + 1, ws.max_row + 1):
            cat = ws.cell(row, col_map["category"]).value
            if not cat or not str(cat).strip():
                continue

            def get(col_key):
                c = col_map.get(col_key)
                if not c:
                    return None
                v = ws.cell(row, c).value
                return v if v is not None and str(v).strip() != "" else None

            def get_num(col_key):
                v = get(col_key)
                if v is None:
                    return None
                try:
                    return float(v)
                except Exception:
                    return None

            def get_int(col_key):
                v = get(col_key)
                if v is None:
                    return None
                try:
                    return int(float(v))
                except Exception:
                    return None

            excel_id_val = get("excel_id")
            excel_id_val = str(excel_id_val).strip() if excel_id_val is not None else None

            data = {
                "excel_id": excel_id_val,
                "category": str(get("category")).strip() if get("category") else None,
                "attr1": str(get("attr1")).strip() if get("attr1") else None,
                "attr2": str(get("attr2")).strip() if get("attr2") else None,
                "attr3": str(get("attr3")).strip() if get("attr3") else None,
                "attr4": str(get("attr4")).strip() if get("attr4") else None,
                "attr5": str(get("attr5")).strip() if get("attr5") else None,
                "brand": str(get("brand")).strip() if get("brand") else None,
                "purchase_price": get_num("purchase_price") or 0,
                "sale_price": get_num("sale_price") or 0,
                "wholesale_price": get_num("wholesale_price") or 0,
                "seller_name": str(get("seller_name")).strip() if get("seller_name") else None,
                "seller_code": str(get("seller_code")).strip() if get("seller_code") else None,
                "purchase_day": get_int("purchase_day"),
                "purchase_month": get_int("purchase_month"),
                "purchase_year": get_int("purchase_year"),
            }

            if excel_id_val:
                existing = await get_product_by_excel_id(excel_id_val)
            else:
                name = build_product_name(
                    data["category"], data["attr1"], data["attr2"],
                    data["attr3"], data["attr4"], data["attr5"],
                    data["brand"], data["seller_name"]
                )
                existing = await get_product_by_name(name)

            excel_barcode = get("barcode")
            excel_stock = get_int("stock")

            if existing:
                data["barcode"] = excel_barcode if excel_barcode is not None else existing.get("barcode")
                data["stock"] = excel_stock if excel_stock is not None else existing.get("stock")
                data["photo_file_id"] = existing.get("photo_file_id")
                data["sale_percent"] = existing.get("sale_percent")
                data["wholesale_percent"] = existing.get("wholesale_percent")
                # اگر تاریخ در اکسل نبود، تاریخ قبلی را نگه دار
                if data["purchase_day"] is None:
                    data["purchase_day"] = existing.get("purchase_day")
                if data["purchase_month"] is None:
                    data["purchase_month"] = existing.get("purchase_month")
                if data["purchase_year"] is None:
                    data["purchase_year"] = existing.get("purchase_year")
                await update_product(existing["id"], data)
                await ensure_initial_history(
                    existing["id"],
                    data.get("purchase_day"), data.get("purchase_month"),
                    data.get("purchase_year"), data.get("purchase_price")
                )
                updated += 1
            else:
                data["barcode"] = str(excel_barcode).strip() if excel_barcode is not None else None
                data["stock"] = excel_stock or 0
                data["photo_file_id"] = None
                data["sale_percent"] = None
                data["wholesale_percent"] = None
                new_id = await add_product(data)
                # add_product خودش هیستوری می‌سازد اگر تاریخ+قیمت باشد
                added += 1

        await state.clear()
        await message.answer(
            "✅ ایمپورت با موفقیت انجام شد.\n"
            f"➕ کالای جدید: {added}\n"
            f"🔄 کالای به‌روزشده: {updated}",
            reply_markup=admin_menu()
        )
    except Exception as e:
        await message.answer(f"❌ خطا در پردازش فایل: {e}")
    finally:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass
